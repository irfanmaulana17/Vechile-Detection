# limit the number of cpus used by high performance libraries
from ast import NotIn
from email.policy import default
from glob import glob
import os
from turtle import distance

from pandas import reset_option
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"

import sys
sys.path.insert(0, './yolov5')

import argparse
import os
import platform
import shutil
import time
from pathlib import Path
import cv2
import numpy as np
import datetime as dtime
from scheduler import Scheduler
import schedule
import csv
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Series, Reference
from openpyxl.styles import *
import scheduler.trigger as trigger
import torch
import torch.backends.cudnn as cudnn

from yolov5.models.experimental import attempt_load
from yolov5.utils.downloads import attempt_download
from yolov5.models.common import DetectMultiBackend
from yolov5.utils.dataloaders import LoadImages, LoadStreams
from yolov5.utils.general import (LOGGER, check_img_size, non_max_suppression, scale_boxes, 
                                  check_imshow, xyxy2xywh, increment_path)
from yolov5.utils.torch_utils import select_device, time_sync
from yolov5.utils.plots import Annotator, colors
from deep_sort.utils.parser import get_config
from deep_sort.deep_sort import DeepSort

FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # yolov5 deepsort root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative
count = 0
s_motor = 0 
s_mobil = 0
s_bus = 0
s_truk = 0
u_motor = 0
u_mobil = 0
u_bus = 0
u_truk = 0
hs_motor = 0 
hs_mobil = 0
hs_bus = 0
hs_truk = 0
hu_motor = 0
hu_mobil = 0
hu_bus = 0
hu_truk = 0
speed_km = 0
row = 2
baris = 2
baris1 = 2
rata = 0
rata1 = 0
head = []
rowsbt = []
rowstb = []
t_motor = [0] 
t_mobil = [0]
t_bus = [0]
t_truk = [0]
b_motor = [0]
b_mobil = [0]
b_bus = [0]
b_truk = [0]
data = []
data_s = []
vechiles_enter1 = {}
vechiles_enterKM1 = {}
vechiles_enter2 = {}
vechiles_enterKM2 = {}
vechiles_elapsed_time = {}
vechiles_elapsed_time2 = {}

def insert ():
    # Membuat workbook baru
    workbook = openpyxl.Workbook()

    # Membuat worksheet baru
    worksheet = workbook.active

    worksheet['A1'] = 'Nama'
    worksheet['A2'] = 'Umur'
    worksheet['A3'] = 'Email'

    workbook.save('laporan.xlsx')

# Fungsi update data
def update_data():
    # Ambil data terbaru
    # ...
    
    # Simpan data ke file Excel
    workbook = openpyxl.load_workbook('laporan.xlsx')
    worksheet = workbook.active
    # ...
    workbook.save('laporan.xlsx')
def detect(opt):
    
    out, source, yolo_model, deep_sort_model, show_vid, save_vid, save_txt, imgsz, evaluate, half, project, name, exist_ok= \
        opt.output, opt.source, opt.yolo_model, opt.deep_sort_model, opt.show_vid, opt.save_vid, \
        opt.save_txt, opt.imgsz, opt.evaluate, opt.half, opt.project, opt.name, opt.exist_ok
    webcam = source == '1' or source.startswith(
        'rtsp') or source.startswith('http') or source.endswith('.txt')

    # initialize deepsort
    cfg = get_config()
    cfg.merge_from_file(opt.config_deepsort)
    deepsort = DeepSort(deep_sort_model,
                        max_dist=cfg.DEEPSORT.MAX_DIST,
                        max_iou_distance=cfg.DEEPSORT.MAX_IOU_DISTANCE,
                        max_age=cfg.DEEPSORT.MAX_AGE, n_init=cfg.DEEPSORT.N_INIT, nn_budget=cfg.DEEPSORT.NN_BUDGET,
                        use_cuda=True)

    # Initialize
    device = select_device(opt.device)
    half &= device.type != 'cpu'  # half precision only supported on CUDA

    # The MOT16 evaluation runs multiple inference streams in parallel, each one writing to
    # its own .txt file. Hence, in that case, the output folder is not restored
    if not evaluate:
        if os.path.exists(out):
            pass
            shutil.rmtree(out)  # delete output folder
        os.makedirs(out)  # make new output folder

    # Directories
    save_dir = increment_path(Path(project) / name, exist_ok=exist_ok)  # increment run
    save_dir.mkdir(parents=True, exist_ok=True)  # make dir

    # Load model
    device = select_device(device)
    model = DetectMultiBackend(yolo_model, device=device, dnn=opt.dnn)
    stride, names, pt, jit, _ = model.stride, model.names, model.pt, model.jit, model.onnx
    imgsz = check_img_size(imgsz, s=stride)  # check image size

    # Half
    half &= pt and device.type != 'cpu'  # half precision only supported by PyTorch on CUDA
    if pt:
        model.model.half() if half else model.model.float()

    # Set Dataloader
    vid_path, vid_writer = None, None
    # Check if environment supports image displays
    if show_vid:
        show_vid = check_imshow()

    # Dataloader
    if webcam:
        show_vid = check_imshow()
        cudnn.benchmark = True  # set True to speed up constant image size inference
        dataset = LoadStreams(source, img_size=imgsz, stride=stride, auto=pt and not jit)
        bs = len(dataset)  # batch_size
    else:
        dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt and not jit)
        bs = 1  # batch_size
    vid_path, vid_writer = [None] * bs, [None] * bs

    # Get names and colors
    names = model.module.names if hasattr(model, 'module') else model.names

    # extract what is in between the last '/' and last '.'
    txt_file_name = source.split('/')[-1].split('.')[0]
    txt_path = str(Path(save_dir)) + '/' + txt_file_name + '.txt'

    if pt and device.type != 'cpu':
        model(torch.zeros(1, 3, *imgsz).to(device).type_as(next(model.model.parameters())))  # warmup
    dt, seen = [0.0, 0.0, 0.0, 0.0], 0
    for frame_idx, (path, img, im0s, vid_cap, s) in enumerate(dataset):
        t1 = time_sync()
        img = torch.from_numpy(img).to(device)
        img = img.half() if half else img.float()  # uint8 to fp16/32
        img /= 255.0  # 0 - 255 to 0.0 - 1.0
        if img.ndimension() == 3:
            img = img.unsqueeze(0)
        t2 = time_sync()
        dt[0] += t2 - t1

        # Inference
        visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if opt.visualize else False
        pred = model(img, augment=opt.augment, visualize=visualize)
        t3 = time_sync()
        dt[1] += t3 - t2

        # Apply NMS
        pred = non_max_suppression(pred, opt.conf_thres, opt.iou_thres, opt.classes, opt.agnostic_nms, max_det=opt.max_det)
        dt[2] += time_sync() - t3

        # Process detections
        for i, det in enumerate(pred):  # detections per image
            seen += 1
            if webcam:  # batch_size >= 1
                p, im0, _ = path[i], im0s[i].copy(), dataset.count
                s += f'{i}: '
            else:
                p, im0, _ = path, im0s.copy(), getattr(dataset, 'frame', 0)

            p = Path(p)  # to Path
            save_path = str(save_dir / p.name)  # im.jpg, vid.mp4, ...
            s += '%gx%g ' % img.shape[2:]  # print string

            annotator = Annotator(im0, line_width=2, pil=not ascii)
           
            w, h = im0.shape[1],im0.shape[0]
            if det is not None and len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_boxes(
                    img.shape[2:], det[:, :4], im0.shape).round()

                # Print results
                for c in det[:, -1].unique():
                    n = (det[:, -1] == c).sum()  # detections per class
                    s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

                xywhs = xyxy2xywh(det[:, 0:4])
                confs = det[:, 4]
                clss = det[:, 5]

                # pass detections to deepsort
                t4 = time_sync()
                outputs = deepsort.update(xywhs.cpu(), confs.cpu(), clss.cpu(), im0)
                t5 = time_sync()
                dt[3] += t5 - t4
                # area


                # draw boxes for visualization
                if len(outputs) > 0:
                    global label
                    for j, (output, conf) in enumerate(zip(outputs, confs)):
                        
                        area1 = [(100,0),(550,0),(820,100),(250,200)]
                        area2 = [(260,230),(870,130),(w-0,280),(550,550)]
                        
                        bboxes = output[0:4]
                        id = output[4]
                        cls = output[5]
                        #count
                        
                        c = int(cls)  # integer class
                        label = f'{id}{names[c]}'
                        count_obj(bboxes,w,h,id)
                        result1 = cv2.pointPolygonTest(np.array(area1, np.int32),(center_coordinates), False)
                        result2 = cv2.pointPolygonTest(np.array(area2, np.int32),(center_coordinates), False)
                        annotator.box_label(bboxes,label, color=colors(c, True))
                        if result1 >= 0:
                            vechiles_enterKM1[id] = time.time()
                        if id in vechiles_enterKM1:
                            if result2 >= 0:
                                if label == 'Plat':
                                    annotator.box_label(bboxes,label, color=colors(c, True))
                                else:
                                    global lbl,hs_motor,hs_mobil,hs_bus,hs_truk,baris
                                    elapsed_time = time.time() - vechiles_enterKM1[id]
                                    if id not in vechiles_elapsed_time:
                                        vechiles_elapsed_time[id] = elapsed_time
                                    if id in vechiles_elapsed_time:
                                        elapsed_time = vechiles_elapsed_time[id]
                                    # Calc Speed
                                    distance = 85 #meter
                                    speed_ms = distance / elapsed_time
                                    speed_km = speed_ms * 3.6
                                    lbl = str(int(speed_km))
                                    annotator.box_label(bboxes,label+' '+lbl+'Km/j', color=colors(c, True))
                                    if label == "Motor":
                                        if id not in data_s:
                                            a = int(lbl)
                                            t_motor.append(a)
                                            data_s.append(id)
                                            hs_motor += 1
                                            if hs_motor not in head :
                                                head.append(hs_motor)
                                            if hs_motor not in rowstb:
                                                baris += 1
                                                rowstb.append(hs_motor)
                                    if label == "Mobil":
                                        if id not in data_s:
                                            a = int(lbl)
                                            t_mobil.append(a)
                                            data_s.append(id)
                                            hs_mobil += 1   
                                            if hs_mobil not in head:
                                                head.append(hs_mobil)
                                            if hs_mobil not in rowstb:
                                                baris += 1
                                                rowstb.append(hs_mobil)
                                    if label == "Bus":
                                        if id not in data_s:
                                            a = int(lbl)
                                            t_bus.append(a)
                                            data_s.append(id)
                                            hs_bus += 1
                                            if hs_bus not in head:
                                                head.append(hs_bus)
                                            if hs_bus not in rowstb:
                                                baris += 1
                                                rowstb.append(hs_bus)
                                    if label == "Truk":
                                        if id not in data_s:
                                            a = int(lbl)
                                            t_truk.append(a)
                                            data_s.append(id)
                                            hs_truk += 1
                                            if hs_truk not in head:
                                                head.append(hs_truk)
                                            if hs_truk not in rowstb:
                                                baris += 1
                                                rowstb.append(hs_truk)
                        if result2 >= 0:
                            vechiles_enterKM2[id] = time.time()
                            annotator.box_label(bboxes,label, color=colors(c, True))
                        if id in vechiles_enterKM2:
                            result1 = cv2.pointPolygonTest(np.array(area1, np.int32), (center_coordinates), False)
                            if result1 >= 0:
                                if label == 'Plat':
                                    annotator.box_label(bboxes,label, color=colors(c, True))
                                else:
                                    global hu_motor,hu_mobil,hu_bus,hu_truk,baris1
                                    elapsed_time2 = time.time() - vechiles_enterKM2[id]
                                    if id not in vechiles_elapsed_time2:
                                        vechiles_elapsed_time2[id] = elapsed_time2
                                    if id in vechiles_elapsed_time2:
                                        elapsed_time2 = vechiles_elapsed_time2[id]
                                    # Calc Speed
                                    distance2 = 100 #meter
                                    speed_ms2 = distance2 / elapsed_time2
                                    speed_km2 = speed_ms2 * 3.6
                                    lbl2 = str(int(speed_km2))
                                    annotator.box_label(bboxes,label+' '+lbl2+'Km/j', color=colors(c, True))
                                    if label == "Motor":
                                        if id not in data_s:
                                            a = int(lbl2)
                                            b_motor.append(a)
                                            data_s.append(id)
                                            hu_motor += 1
                                            if hu_motor not in head:
                                                head.append(hu_motor)
                                            if hu_motor not in rowsbt:
                                                baris1 += 1
                                                rowsbt.append(hu_motor)
                                
                                    if label == "Mobil":
                                        if id not in data_s:
                                            a = int(lbl2)
                                            b_mobil.append(a)
                                            data_s.append(id)
                                            hu_mobil += 1
                                            if hu_mobil not in head:
                                                head.append(hu_mobil)
                                            if hu_mobil not in rowsbt:
                                                baris1 += 1
                                                rowsbt.append(hu_mobil)
                                    if label == "Bus":
                                        if id not in data_s:
                                            a = int(lbl2)
                                            b_bus.append(a)
                                            data_s.append(id)
                                            hu_bus += 1
                                            if hu_bus not in head:
                                                head.append(hu_bus)
                                            if hu_bus not in rowsbt:
                                                baris1 += 1
                                                rowsbt.append(hu_bus)
                                    if label == "Truk":
                                        if id not in data_s:
                                            a = int(lbl2)
                                            b_truk.append(a)
                                            data_s.append(id)
                                            hu_truk += 1
                                            if hu_truk not in head:
                                                head.append(hu_truk)
                                            if hu_truk not in rowsbt:
                                                baris1 += 1
                                                rowsbt.append(hu_truk)
                        if save_txt:
                            # to MOT format
                            bbox_left = output[0]
                            bbox_top = output[1]
                            bbox_w = output[2] - output[0]
                            bbox_h = output[3] - output[1]
                            # Write MOT compliant results to file
                            with open(txt_path, 'a') as f:
                                f.write(('%g ' * 10 + '\n') % (frame_idx + 1, id, bbox_left,  # MOT format
                                                            bbox_top, bbox_w, bbox_h, -1, -1, -1, -1))
                        
                LOGGER.info(f'{s}Done. YOLO:({t3 - t2:.3f}s), DeepSort:({t5 - t4:.3f}s)')

            else:
                deepsort.increment_ages()
                LOGGER.info('No detections')

            # Stream results
            im0 = annotator.result()
            if show_vid:
                # global count
                color_font = (0,215,255)
                custom_color = (0,255,0)
                im0_res = cv2.resize(im0, (854, 480))

                area1 = [(0,0),(950,0),(880,300),(0,300)]
                area2 = [(0,425),(870,425),(780,1080),(0,1080)]

                for area in [area1,area2]:
                    cv2.polylines(im0,[np.array(area, np.int32)], True,(custom_color),2)
                
                # thickness = 2
                # font = cv2.FONT_HERSHEY_DUPLEX
                # fontScale = 0.75

                # cv2.putText(im0,' Dari Barat', (50, 20), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
                # cv2.putText(im0,' Dari Timur', (1050, 20), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
                # cv2.putText(im0, str(u_motor) + ' Sepeda Motor', (50, 50), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
                # cv2.putText(im0, str(u_mobil) + ' Mobil', (50, 80), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
                # cv2.putText(im0, str(u_bus) + ' Bus', (50, 110), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
                # cv2.putText(im0, str(u_truk) + ' Truk', (50,140), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
                # cv2.putText(im0, str(s_motor) + ' Sepeda Motor', (1050, 50), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
                # cv2.putText(im0, str(s_mobil) + ' Mobil', (1050, 80), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
                # cv2.putText(im0, str(s_bus) + ' Bus', (1050, 110), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
                # cv2.putText(im0, str(s_truk) + ' Truk', (1050,140), font, 
                #     fontScale, color_font, thickness, cv2.LINE_AA)
               
                cv2.imshow(str(p), im0)
                if cv2.waitKey(1) == ord('q'):  # q to quit
                    raise StopIteration

            # Save results (image with detections)
            if save_vid:
                if vid_path != save_path:  # new video
                    vid_path = save_path
                    if isinstance(vid_writer, cv2.VideoWriter):
                        vid_writer.release()  # release previous video writer
                    if vid_cap:  # video
                        fps = vid_cap.get(cv2.CAP_PROP_FPS)
                        w = int(vid_cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                        h = int(vid_cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                    else:  # stream
                        fps, w, h = 30, im0.shape[1], im0.shape[0]

                    vid_writer = cv2.VideoWriter(save_path, cv2.VideoWriter_fourcc(*'mp4v'), fps, (w, h))
                vid_writer.write(im0)

    # Print results
    t = tuple(x / seen * 1E3 for x in dt)  # speeds per image
    LOGGER.info(f'Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS, %.1fms deep sort update \
        per image at shape {(1, 3, *imgsz)}' % t)
    if save_txt or save_vid:
        print('Results saved to %s' % save_path)
        if platform == 'darwin':  # MacOS
            os.system('open ' + save_path)

def count_obj(box,w,h,id):
    global count,data,center_coordinates
    global s_motor,s_mobil,s_bus,s_truk,u_motor,u_mobil,u_bus,u_truk,label
    waktu = dtime.datetime.now()
    center_coordinates = (int(box[0]+(box[2]-box[0])/2) , int(box[1]+(box[3]-box[1])/2))
    area1 = [(100,0),(550,0),(820,100),(250,200)]
    area2 = [(260,230),(870,130),(w-0,280),(550,550)]
    result1 = cv2.pointPolygonTest(np.array(area1, np.int32), (center_coordinates), False)
    result2 = cv2.pointPolygonTest(np.array(area2, np.int32), (center_coordinates), False)
    
    
    if result1 >= 0:
       vechiles_enter1[id] = center_coordinates
    if id in vechiles_enter1:
        if result2 >= 0:  
            if label == "Mobil":
                if  id not in data:
                    s_mobil += 1
                    data.append(id)
            if label == "Motor":
                if  id not in data:
                    s_motor += 1
                    data.append(id)
            if label == "Bus":
                if  id not in data:
                    s_bus += 1
                    data.append(id)
            if label == "Truk":
                if  id not in data:
                    s_truk += 1
                    data.append(id)
    if result2 >= 0:
       vechiles_enter2[id] = center_coordinates    
    if id in vechiles_enter2:
        if result1 >= 0:  
            if label == "Mobil":
                if  id not in data:
                    u_mobil += 1
                    data.append(id)
            if label == "Motor":
                if  id not in data:
                    u_motor += 1
                    data.append(id)
            if label == "Bus":
                if  id not in data:
                    u_bus += 1
                    data.append(id)
            if label == "Truk":
                if  id not in data:
                    u_truk += 1
                    data.append(id) 
if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--yolo_model', nargs='+', type=str, default='new.pt', help='model.pt path(s)')
    parser.add_argument('--deep_sort_model', type=str, default='osnet_x0_25')
    parser.add_argument('--source', type=str, default='rtsp://admin:admin@192.168.35.6:8554/Streaming/Channels/101', help='source')  # file/folder, 0 for webcam
    parser.add_argument('--output', type=str, default='inference/output', help='output folder')  # output folder
    parser.add_argument('--imgsz', '--img', '--img-size', nargs='+', type=int, default= [640], help='inference size h,w')
    parser.add_argument('--conf-thres', type=float, default=0.5, help='object confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.5, help='IOU threshold for NMS')
    parser.add_argument('--fourcc', type=str, default='mp4v', help='output video codec (verify ffmpeg support)')
    parser.add_argument('--device', default='', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--show-vid', action='store_false', help='display tracking video results')
    parser.add_argument('--save-vid', action='store_true',default='--save-vid', help='save video tracking results')
    parser.add_argument('--save-txt', action='store_true', help='save MOT compliant results to *.txt')
    # class 0 is person, 1 is bycicle, 2 is car... 79 is oven
    parser.add_argument('--classes', nargs='+', type=int, help='filter by class: --class 0, or --class 16 17')
    parser.add_argument('--agnostic-nms', action='store_true', help='class-agnostic NMS')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--evaluate', action='store_true', help='augmented inference')
    parser.add_argument("--config_deepsort", type=str, default="deep_sort/configs/deep_sort.yaml")
    parser.add_argument("--half", action="store_true", help="use FP16 half-precision inference")
    parser.add_argument('--visualize', action='store_true', help='visualize features')
    parser.add_argument('--max-det', type=int, default=1000, help='maximum detection per image')
    parser.add_argument('--dnn', action='store_true', help='use OpenCV DNN for ONNX inference')
    parser.add_argument('--project', default=ROOT / 'runs/track', help='save results to project/name')
    parser.add_argument('--name', default='exp', help='save results to project/name')
    parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
    opt = parser.parse_args()
    opt.imgsz *= 2 if len(opt.imgsz) == 1 else 1  # expand

    with torch.no_grad():
        detect(opt)
