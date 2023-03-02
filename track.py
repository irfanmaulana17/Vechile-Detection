# limit the number of cpus used by high performance libraries
import os
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"

import sys
sys.path.insert(0, './yolov5')

import argparse
import os
import platform
import shutil
import time
from pathlib import Path
import cv2
import numpy as np
import torch
import torch.backends.cudnn as cudnn
import tkinter as tk
from tkinter import *
from tkinter import ttk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
import pytesseract
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
from datetime import datetime
import schedule
import PIL
from PIL import Image,ImageTk

from yolov5.models.experimental import attempt_load
from yolov5.utils.downloads import attempt_download
from yolov5.models.common import DetectMultiBackend
from yolov5.utils.dataloaders import LoadImages, LoadStreams
from yolov5.utils.general import (LOGGER, check_img_size, non_max_suppression, scale_boxes, 
                                  check_imshow, xyxy2xywh, increment_path)
from yolov5.utils.torch_utils import select_device, time_sync
from yolov5.utils.plots import Annotator, colors
from deep_sort.utils.parser import get_config
from deep_sort.deep_sort import DeepSort

FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # yolov5 deepsort root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative
count = 0
data = []
data_s = []
data_p = []
b_motor = 0 
b_mobil = 0
b_bus = 0
b_truk = 0
t_motor = 0 
t_mobil = 0
t_bus = 0
t_truk = 0
speed_km = 0
row = 2
baris = 2
rata = 0
rata1 = 0
motorcycle_speeds = []
data_kendaraan = []
head = []
rowsbt = []
rowstb = []
vechiles_enter1 = {}
vechiles_enterKM1 = {}
vechiles_enter2 = {}
vechiles_enterKM2 = {}
vechiles_elapsed_time = {}
vechiles_elapsed_time2 = {}
def reset():
    global b_motor,b_mobil,b_bus,b_truk
    b_motor = 0 
    b_mobil = 0
    b_bus = 0
    b_truk = 0
    # panggil fungsi reset ini lagi setelah 1 menit
    windows.after(15000, reset)
def show_plot_in_tkinter(fig, master):
    canvas = FigureCanvasTkAgg(fig, master=master)
    canvas.draw()
    canvas.get_tk_widget().grid(column=1,row=0,padx=0,pady=0)
def create_plot(files):
    fig, ax = plt.subplots(figsize=(6, 4))
    files['Waktu'] = pd.to_datetime(files['Waktu'], format='%H:%M:%S')
    ax.plot(files['Waktu'], files['Volume'])
    ax.set_xlabel('Data')
    ax.set_ylabel('Value')
    ax.tick_params(axis='x', labelrotation=90,labelsize=9)
    return fig
def update():
    labelmt.config(text=str(b_motor))
    labelmb.config(text=str(b_mobil))
    labelbs.config(text=str(b_bus))
    labeltk.config(text=str(b_truk))
    windows.after(500,update)

def tkinter():
    global ca,windows,label_frame_dt
    global labelmt,labelmb,labelbs,labeltk
    windows = tk.Tk()
    title = "Aplikasi Counting Vechile"
    geometri = "760x500"
    windows.geometry(geometri)
    windows.title(title)
    
    # FaceAbsen = tk.Button(windows, font=("Arial", 18), text="Camera 1", bg="blue", fg="white", height=3, width=25, command=detect).grid(row=2,column=0, padx=10,pady=10)
    # FaceAbsen = tk.Button(windows, font=("Arial", 18), text="Camera 2", bg="blue", fg="white", height=3, width=25, command=detect1).grid(row=2,column=2, padx=10,pady=10)
    label_frame = LabelFrame(windows, text='Home Program')
    label_frame.grid(column=0,row=0,padx=10,pady=10,sticky="w")
    label_frame_vd = LabelFrame(windows, text='Tampilan Layar')
    label_frame_vd.place(x=350, y=10, width = 904, height = 530)
    label_frame_dt = LabelFrame(windows, text='Tampilan Data')
    label_frame_dt.grid(column=0,row=1, padx= 10 ,pady= 380)
    button1 = tk.Button(label_frame, text = "Start Program",width=30,command=detect)
    button1.grid(row = 0, column = 0, padx = 25, pady = 10,columnspan=4)
    # button1 = tk.Button(label_frame, text = "Open History",width=30,command=laporan)
    # button1.grid(row = 3, column = 0, padx = 25, pady = 10,columnspan=4)
    # lbl = tk.Label(label_frame_vd,text='Home Program')
    ca = tk.Canvas(label_frame_vd, width=854, height=480, bg='white')
    ca.place(x=10, y=10)
    label = Label(label_frame, text='Motor')
    label.grid (column = 0,row = 1, padx = 10, pady = 10)
    label1 = Label(label_frame, text='Mobil')
    label1.grid (column = 1,row = 1, padx = 10, pady = 10)
    label2 = Label(label_frame, text='Bus')
    label2.grid (column = 2,row = 1, padx = 10, pady = 10)
    label3 = Label(label_frame, text='Truk')
    label3.grid (column = 3,row = 1, padx = 10, pady = 10)
    labelmt = Label(label_frame)
    labelmt.grid (column = 0,row = 2, padx = 10, pady = 10)
    labelmb = Label(label_frame)
    labelmb.grid (column = 1,row = 2, padx = 10, pady = 10)
    labelbs = Label(label_frame)
    labelbs.grid (column = 2,row = 2, padx = 10, pady = 10)
    labeltk = Label(label_frame)
    labeltk.grid (column = 3,row = 2, padx = 10, pady = 10)
    # ca1 = Canvas(windows, width=1000, height=800, bg='white')
    # ca1.place(x=10, y=10, width = 1000, height = 800)
    # detect()
    save_data()
    new_file()
    reset()
    # fig = create_plot(files)
    # show_plot_in_tkinter(fig,label_frame_dt)
    load_data()
    load_data_kecepatan()
    load_pivot()
    update()
    # detect()
    windows.mainloop()
def save_data():
    filename = 'Data.xlsx'
    # get the current time
    now = datetime.now()
    total = b_motor+b_mobil+b_bus+b_truk
    # format the current time as a string
    current_time = now.strftime("%H:%M:%S")
# cek apakah file excel sudah ada atau belum
    if os.path.isfile(filename):
        # jika file excel sudah ada, buka file tersebut dan tambahkan data baru ke sheet yang ada
        workbook = load_workbook(filename)
        worksheet = workbook.active
        # tambahkan data ke file excel
        # misalnya dengan menambahkan data pada baris baru setelah baris terakhir yang terisi
        last_row = worksheet.max_row
        worksheet.cell(row=last_row+1, column=1).value = current_time
        # worksheet.cell(row=last_row+1, column=2).value = 'Data 2'
        worksheet.cell(row=last_row+1, column=2).value = total
        # worksheet.cell(row=last_row+1, column=4).value = 'Data 2'
        worksheet.cell(row=last_row+1, column=3).value = b_motor
        worksheet.cell(row=last_row+1, column=4).value = b_mobil
        worksheet.cell(row=last_row+1, column=5).value = b_bus
        worksheet.cell(row=last_row+1, column=6).value = b_truk
        # worksheet.cell(row=last_row+1, column=9).value = 'Data 1'
        # worksheet.cell(row=last_row+1, column=10).value = 'Data 2'
        # worksheet.cell(row=last_row+1, column=11).value = 'Data 1'
        # worksheet.cell(row=last_row+1, column=12).value = 'Data 2'
        workbook.save(filename)
    else:
        # jika file excel belum ada, buat file baru dan tambahkan data ke sheet baru
        workbook = Workbook()
        worksheet = workbook.active
        # buat header untuk file excel
        worksheet.cell(row=1, column=1).value = 'Waktu'
        worksheet.cell(row=1, column=2).value = 'Barat'
        # worksheet.cell(row=1, column=2).value = 'Timur'
        # worksheet.cell(row=1, column=4).value = 'Dua Arah'
        worksheet.cell(row=1, column=3).value = 'Motor Barat'
        worksheet.cell(row=1, column=4).value = 'Mobil Barat'
        worksheet.cell(row=1, column=5).value = 'Bus Barat'
        worksheet.cell(row=1, column=6).value = 'Truk Barat'
        # worksheet.cell(row=1, column=9).value = 'Motor Timur'
        # worksheet.cell(row=1, column=10).value = 'Mobil Timur'
        # worksheet.cell(row=1, column=11).value = 'Bus Timur'
        # worksheet.cell(row=1, column=12).value = 'Truk Timur'
        # tambahkan data ke file excel
        # last_row = worksheet.max_row
        worksheet.cell(row=2, column=1).value = current_time
        # worksheet.cell(row=2, column=2).value = 'Data 2'
        worksheet.cell(row=2, column=2).value = total
        # worksheet.cell(row=2, column=4).value = 'Data 2'
        worksheet.cell(row=2, column=3).value = t_motor
        worksheet.cell(row=2, column=4).value = t_mobil
        worksheet.cell(row=2, column=5).value = t_bus
        worksheet.cell(row=2, column=6).value = t_truk
        # worksheet.cell(row=2, column=9).value = 'Data 1'
        # worksheet.cell(row=2, column=10).value = 'Data 2'
        # worksheet.cell(row=2, column=11).value = 'Data 1'
        # worksheet.cell(row=2, column=12).value = 'Data 2'
        workbook.save(filename)
    windows.after(15000,save_data)
def load_data():
    global files
    files = pd.read_excel("Data.xlsx")
    # files = files.astype(int)
    # files['Waktu'] = files['Waktu'].apply(lambda x: datetime.strftime(x, '%H:%M') if isinstance(x, datetime) else x)
    # files['Arah Barat'] = files['Arah Barat'].astype(int)s
    header = list(files.columns)
    tree = ttk.Treeview(label_frame_dt, columns=header, show="headings")
    tree.grid(column=0,row=0,padx=10,pady=10,sticky="wn")

    for col in header:
        tree.column(col, width=100)
        tree.heading(col,text=col)

    for index, row in files.iterrows():
        values = list(row.values)  
        tree.insert("", index, values=values)

    windows.after(500,load_data)
def save_data_kecepatan():
    filename = 'Data Kecepatan.xlsx'
    # get the current time
    now = datetime.now()

    # format the current time as a string
    current_time = now.strftime("%H:%M:%S")


    # jika file excel sudah ada, buka file tersebut dan tambahkan data baru ke sheet yang ada
    workbook = load_workbook(filename)
    worksheet = workbook.active
    # tambahkan data ke file excel
    # misalnya dengan menambahkan data pada baris baru setelah baris terakhir yang terisi
    last_row = worksheet.max_row
    worksheet.cell(row=last_row+1, column=1).value = current_time
    worksheet.cell(row=last_row+1, column=2).value = data_kendaraan[-1]
    worksheet.cell(row=last_row+1, column=3).value = motorcycle_speeds[-1]
    workbook.save(filename)
def new_file():
    filename = 'Data Kecepatan.xlsx'
    if os.path.isfile(filename):
        workbook = load_workbook(filename)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        # buat header untuk file excel
        worksheet.cell(row=1, column=1).value = 'Waktu'
        worksheet.cell(row=1, column=2).value = 'Jenis Kendaraan'
        worksheet.cell(row=1, column=3).value = 'Kecepatan Km/h'
        workbook.save(filename)
    file = 'Data Kecepatan Pivot.xlsx'
    if os.path.isfile(file):
        workbook = load_workbook(file)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        # buat header untuk file excel
        worksheet.cell(row=1, column=1).value = 'Jenis Kendaraan'
        worksheet.cell(row=1, column=2).value = 'Rata-rata Km/h'
        workbook.save(file)
def load_data_kecepatan():
    global files
    files = pd.read_excel("Data Kecepatan.xlsx")
    # files = files.astype(int)
    # files['Waktu'] = files['Waktu'].apply(lambda x: datetime.strftime(x, '%H:%M') if isinstance(x, datetime) else x)
    # files['Arah Barat'] = files['Arah Barat'].astype(int)s
    header = list(files.columns)
    tree = ttk.Treeview(label_frame_dt, columns=header, show="headings")
    tree.grid(column=2,row=0,padx=10,pady=10,sticky="wn")

    for col in header:
        tree.column(col, width=120)
        tree.heading(col,text=col)

    for index, row in files.iterrows():
        values = list(row.values)  
        tree.insert("", index, values=values)

    windows.after(500,load_data_kecepatan)

def pivot():
# Membaca data dari file Excel
    df = pd.read_excel('Data Kecepatan.xlsx')

    # Melakukan pivoting data berdasarkan kolom "Jenis Kendaraan"
    df_pivot = pd.pivot_table(df, values='Kecepatan Km/h', index='Jenis Kendaraan', aggfunc='mean')
    df_pivot.rename(columns={'Kecepatan Km/h': 'Rata-rata Km/h'}, inplace=True)
    # Menyimpan data hasil pivoting ke dalam file Excel
    df_pivot.to_excel('Data Kecepatan Pivot.xlsx')

def load_pivot():
    files = pd.read_excel("Data Kecepatan Pivot.xlsx")
    # files = files.astype(int)
    # files['Waktu'] = files['Waktu'].apply(lambda x: datetime.strftime(x, '%H:%M') if isinstance(x, datetime) else x)
    # files['Arah Barat'] = files['Arah Barat'].astype(int)s
    header = list(files.columns)
    tree = ttk.Treeview(label_frame_dt, columns=header, show="headings")
    tree.grid(column=3,row=0,padx=10,pady=10,sticky="wn")

    for col in header:
        tree.column(col, width=100)
        tree.heading(col,text=col)

    for index, row in files.iterrows():
        values = list(row.values)  
        tree.insert("", index, values=values)

    windows.after(500,load_pivot)
def detect():
    parser = argparse.ArgumentParser()
    parser.add_argument('--yolo_model', nargs='+', type=str, default='dataset/best.pt', help='model.pt path(s)')
    parser.add_argument('--deep_sort_model', type=str, default='osnet_x0_25')
    parser.add_argument('--source', type=str, default='videos/IMG_0693.MOV', help='source')  # file/folder, 0 for webcam
    parser.add_argument('--output', type=str, default='inference/output', help='output folder')  # output folder
    parser.add_argument('--imgsz', '--img', '--img-size', nargs='+', type=int, default=[480], help='inference size h,w')
    parser.add_argument('--conf-thres', type=float, default=0.5, help='object confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.5, help='IOU threshold for NMS')
    parser.add_argument('--fourcc', type=str, default='mp4v', help='output video codec (verify ffmpeg support)')
    parser.add_argument('--device', default='', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--show-vid', action='store_false', help='display tracking video results')
    parser.add_argument('--save-vid', action='store_true', help='save video tracking results')
    parser.add_argument('--save-txt', action='store_true', help='save MOT compliant results to *.txt')
    # class 0 is person, 1 is bycicle, 2 is car... 79 is oven
    parser.add_argument('--classes', nargs='+', type=int, help='filter by class: --class 0, or --class 16 17')
    parser.add_argument('--agnostic-nms', action='store_true', help='class-agnostic NMS')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--evaluate', action='store_true', help='augmented inference')
    parser.add_argument("--config_deepsort", type=str, default="deep_sort/configs/deep_sort.yaml")
    parser.add_argument("--half", action="store_true", help="use FP16 half-precision inference")
    parser.add_argument('--visualize', action='store_true', help='visualize features')
    parser.add_argument('--max-det', type=int, default=1000, help='maximum detection per image')
    parser.add_argument('--dnn', action='store_true', help='use OpenCV DNN for ONNX inference')
    parser.add_argument('--project', default=ROOT / 'runs/track', help='save results to project/name')
    parser.add_argument('--name', default='exp', help='save results to project/name')
    parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
    opt = parser.parse_args()
    opt.imgsz *= 2 if len(opt.imgsz) == 1 else 1  # expand
    out, source, yolo_model, deep_sort_model, show_vid, save_vid, save_txt, imgsz, evaluate, half, project, name, exist_ok= \
        opt.output, opt.source, opt.yolo_model, opt.deep_sort_model, opt.show_vid, opt.save_vid, \
        opt.save_txt, opt.imgsz, opt.evaluate, opt.half, opt.project, opt.name, opt.exist_ok
    webcam = source == '0' or source.startswith(
        'rtsp') or source.startswith('http') or source.endswith('.streams')

    # cameras = []
    # for sources in source:
    #     cap = cv2.VideoCapture(sources)
    #     cameras.append(cap)
    # print(cameras)
    # initialize deepsort
    cfg = get_config()
    cfg.merge_from_file(opt.config_deepsort)
    deepsort = DeepSort(deep_sort_model,
                        max_dist=cfg.DEEPSORT.MAX_DIST,
                        max_iou_distance=cfg.DEEPSORT.MAX_IOU_DISTANCE,
                        max_age=cfg.DEEPSORT.MAX_AGE, n_init=cfg.DEEPSORT.N_INIT, nn_budget=cfg.DEEPSORT.NN_BUDGET,
                        use_cuda=True)

    # Initialize
    device = select_device(opt.device)
    half &= device.type != 'cpu'  # half precision only supported on CUDA

    # The MOT16 evaluation runs multiple inference streams in parallel, each one writing to
    # its own .txt file. Hence, in that case, the output folder is not restored
    if not evaluate:
        if os.path.exists(out):
            pass
            shutil.rmtree(out)  # delete output folder
        os.makedirs(out)  # make new output folder

    # Directories
    save_dir = increment_path(Path(project) / name, exist_ok=exist_ok)  # increment run
    save_dir.mkdir(parents=True, exist_ok=True)  # make dir

    # Load model
    device = select_device(device)
    model = DetectMultiBackend(yolo_model, device=device, dnn=opt.dnn)
    stride, names, pt, jit, _ = model.stride, model.names, model.pt, model.jit, model.onnx
    imgsz = check_img_size(imgsz, s=stride)  # check image size

    # Half
    half &= pt and device.type != 'cpu'  # half precision only supported by PyTorch on CUDA
    if pt:
        model.model.half() if half else model.model.float()

    # Set Dataloader
    vid_path, vid_writer = None, None
    # Check if environment supports image displays
    if show_vid:
        show_vid = check_imshow()

    # Dataloader
    if webcam:
        show_vid = check_imshow()
        cudnn.benchmark = True  # set True to speed up constant image size inference
        dataset = LoadStreams(source, img_size=imgsz, stride=stride, auto=pt and not jit)
        bs = len(dataset)  # batch_size
        print(source)
    else:
        dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt and not jit)
        bs = 1  # batch_size
    vid_path, vid_writer = [None] * bs, [None] * bs

    # Get names and colors
    names = model.module.names if hasattr(model, 'module') else model.names

    # extract what is in between the last '/' and last '.'
    txt_file_name = source.split('/')[-1].split('.')[0]
    txt_path = str(Path(save_dir)) + '/' + txt_file_name + '.txt'

    if pt and device.type != 'cpu':
        model(torch.zeros(1, 3, *imgsz).to(device).type_as(next(model.model.parameters())))  # warmup
    dt, seen = [0.0, 0.0, 0.0, 0.0], 0
    for frame_idx, (path, img, im0s, vid_cap, s) in enumerate(dataset):
        t1 = time_sync()
        img = torch.from_numpy(img).to(device)
        img = img.half() if half else img.float()  # uint8 to fp16/32
        img /= 255.0  # 0 - 255 to 0.0 - 1.0
        if img.ndimension() == 3:
            img = img.unsqueeze(0)
        t2 = time_sync()
        dt[0] += t2 - t1

        # Inference
        visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if opt.visualize else False
        pred = model(img, augment=opt.augment, visualize=visualize)
        t3 = time_sync()
        dt[1] += t3 - t2

        # Apply NMS
        pred = non_max_suppression(pred, opt.conf_thres, opt.iou_thres, opt.classes, opt.agnostic_nms, max_det=opt.max_det)
        dt[2] += time_sync() - t3


        for i, det in enumerate(pred):  # detections per image
            seen += 1
            if webcam:  # batch_size >= 1
                p, im0, _ = path[i], im0s[i].copy(), dataset.count
                s += f'{i}: '
            else:
                p, im0, _ = path, im0s.copy(), getattr(dataset, 'frame', 0)

            p = Path(p)  # to Path
            save_path = str(save_dir / p.name)  # im.jpg, vid.mp4, ...
            s += '%gx%g ' % img.shape[2:]  # print string

            annotator = Annotator(im0, line_width=2, pil=not ascii)
            w, h = im0.shape[1],im0.shape[0]
            if det is not None and len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_boxes(
                    img.shape[2:], det[:, :4], im0.shape).round()

                # Print results
                for c in det[:, -1].unique():
                    n = (det[:, -1] == c).sum()  # detections per class
                    s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

                xywhs = xyxy2xywh(det[:, 0:4])
                confs = det[:, 4]
                clss = det[:, 5]

                # pass detections to deepsort
                t4 = time_sync()
                outputs = deepsort.update(xywhs.cpu(), confs.cpu(), clss.cpu(), im0)
                t5 = time_sync()
                dt[3] += t5 - t4

                # draw boxes for visualization
                if len(outputs) > 0:
                    for j, (output, conf) in enumerate(zip(outputs, confs)):
                        global label
                        bboxes = output[0:4]
                        id = output[4]
                        cls = output[5]
                        #count
                        c = int(cls)  # integer class
                        label = f'{names[c]}'
                        count_obj(bboxes,w,h,id)
                        area1 = [(800, 260), (1190, 260), (1270, 490), (760, 490)]
                        area2 = [(742, 575), (1295, 575), (1450, 1020), (650, 1020)] #650
                        area3 = [(688, 855), (1395, 855), (1450, 1020), (650, 1020)] #650
                        result1 = cv2.pointPolygonTest(np.array(area1, np.int32),(center_coordinates), False)
                        result2 = cv2.pointPolygonTest(np.array(area2, np.int32),(center_coordinates), False)
                        result3 = cv2.pointPolygonTest(np.array(area3, np.int32),(center_coordinates), False)
                        # annotator.box_label(bboxes, label, color=colors(c, True))
                        if result1 >= 0:
                            vechiles_enterKM1[id] = time.time()
                            annotator.box_label(bboxes, label, color=colors(c, True))
                        if id in vechiles_enterKM1:
                            if result2 >= 0:
                                if label != 'Plat':
                                    global speed_km
                                    elapsed_time = time.time() - vechiles_enterKM1[id]
                                    if id not in vechiles_elapsed_time:
                                        vechiles_elapsed_time[id] = elapsed_time
                                    if id in vechiles_elapsed_time:
                                        elapsed_time = vechiles_elapsed_time[id]
                                    # Calc Speed    
                                    distance = 85 #meter
                                    speed_ms = distance / elapsed_time
                                    speed_km = speed_ms * 3.6
                                    lbl = str(int(speed_km))
                                    annotator.box_label(bboxes,label+' '+lbl+'Km/j', color=colors(c, True))
                                    if  id not in data_s:
                                        data_s.append(id)
                                        motorcycle_speeds.append(int(lbl))
                                        data_kendaraan.append(str(label))
                                        save_data_kecepatan()
                                        pivot()
                                    motorcycle_speeds.clear()
                                    data_kendaraan.clear()
                        if result3 >= 0:
                            if label == 'Plat':
                                if id not in data_p:
                                    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
                                    crop = im0[int(bboxes[1]):int(bboxes[3]), int(bboxes[0]):int(bboxes[2])]
                                    sharp_kernel = np.array([[-1, -1, -1],
                                                            [0, 7, 0],
                                                            [-1, -1, -1]])
                                    img_gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                                    smoth = cv2.bilateralFilter(img_gray, 10,17,17)
                                    sharp_img = cv2.filter2D(src=smoth, ddepth=-1, kernel=sharp_kernel)
                                    noise = cv2.fastNlMeansDenoising(sharp_img, None, 20, 7, 21)
                                    # create directory to store cropped images if not exists
                                    # Apply OCR
                                    plate_number = pytesseract.image_to_string(noise, config='-l eng --oem 3 --psm 7 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789')
                                    print(plate_number)
                                    if not os.path.exists('cropped_plates'):
                                        os.makedirs('cropped_plates')
                                    # generate unique filename for cropped image
                                    filename = '{}_{}.png'.format(plate_number, int(time.time()))
                                    # save cropped image to folder
                                    cv2.imwrite(os.path.join('cropped_plates', filename), noise)
                                    data_p.append(id)
                                    annotator.box_label(bboxes,plate_number, color=colors(c, True))
                                if id in data_p:
                                    annotator.box_label(bboxes,plate_number, color=colors(c, True))
                        if save_txt:
                            # to MOT format
                            bbox_left = output[0]
                            bbox_top = output[1]
                            bbox_w = output[2] - output[0]
                            bbox_h = output[3] - output[1]
                            # Write MOT compliant results to file
                            with open(txt_path, 'a') as f:
                                f.write(('%g ' * 10 + '\n') % (frame_idx + 1, id, bbox_left,  # MOT format
                                                            bbox_top, bbox_w, bbox_h, -1, -1, -1, -1))

                LOGGER.info(f'{s}Done. YOLO:({t3 - t2:.3f}s), DeepSort:({t5 - t4:.3f}s)')

            else:
                deepsort.increment_ages()
                LOGGER.info('No detections')

            # Stream results
            # print("Before:", im0)
            im0 = annotator.result()
            if show_vid:
                # print("After:", im0) 
                im0_resize = cv2.resize(im0, (854, 480))   
                global count
                custom_color = (0,255,0)
                area1 = [(357, 115), (530, 115), (565, 220), (338, 220)]
                area2 = [(332, 255), (575, 255), (645, 455), (290, 455)] #290
                area3 = [(305, 380), (620, 380), (645, 455), (290, 455)] #290
                for area in [area1,area2,area3]:
                    cv2.polylines(im0_resize,[np.array(area, np.int32)], True,(custom_color),2)
                global photo
                photo = ImageTk.PhotoImage(image = Image.fromarray(im0_resize[:, :, ::-1]))
                # Add a PhotoImage to the Canvas
                ca.create_image(0, 0, image=photo, anchor=tk.NW)
                windows.update()
                # cv2.imshow(str(p), im0)
                if cv2.waitKey(1) == ord('q'):  # q to quit
                    raise StopIteration

            # Save results (image with detections)
            if save_vid:
                if vid_path != save_path:  # new video
                    vid_path = save_path
                    if isinstance(vid_writer, cv2.VideoWriter):
                        vid_writer.release()  # release previous video writer
                    if vid_cap:  # video
                        fps = vid_cap.get(cv2.CAP_PROP_FPS)
                        w = int(vid_cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                        h = int(vid_cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                    else:  # stream
                        fps, w, h = 30, im0.shape[1], im0.shape[0]

                    vid_writer = cv2.VideoWriter(save_path, cv2.VideoWriter_fourcc(*'mp4v'), fps, (w, h))
                vid_writer.write(im0)

    # Print results
    t = tuple(x / seen * 1E3 for x in dt)  # speeds per image
    LOGGER.info(f'Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS, %.1fms deep sort update \
        per image at shape {(1, 3, *imgsz)}' % t)
    if save_txt or save_vid:
        print('Results saved to %s' % save_path)
        if platform == 'darwin':  # MacOS
            os.system('open ' + save_path)

def count_obj(box,w,h,id):
    global count,data,center_coordinates
    global b_motor,b_mobil,b_bus,b_truk
    center_coordinates = (int(box[0]+(box[2]-box[0])/2) , int(box[1]+(box[3]-box[1])/2))
    area1 = [(800, 260), (1190, 260), (1270, 490), (760, 490)]
    area2 = [(742, 575), (1295, 575), (1450, 1020), (650, 1020)] #650
    result1 = cv2.pointPolygonTest(np.array(area1, np.int32),(center_coordinates), False)
    result2 = cv2.pointPolygonTest(np.array(area2, np.int32),(center_coordinates), False)
    if result1 >= 0:
       vechiles_enter1[id] = center_coordinates
    if id in vechiles_enter1:
        if result2 >= 0:
            if label == "Mobil":
                if  id not in data:
                    b_mobil += 1
                    data.append(id)
            if label == "Motor":
                if  id not in data:
                    b_motor += 1
                    data.append(id)
            if label == "Bus":
                if  id not in data:
                    b_bus += 1
                    data.append(id)
            if label == "Truk":
                if  id not in data:
                    b_truk += 1
                    data.append(id)


if __name__ == '__main__':


    with torch.no_grad():
        tkinter()